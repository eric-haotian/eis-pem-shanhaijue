# learning AI at www.haotianblog.com
"""CSV ingestion and XLSX export for the desktop interface.

The reader is deliberately permissive: any impedance spectrum should be identifiable, so
the only genuinely required information is frequency and complex impedance. Everything
else is inferred, defaulted, or reconstructed, and every inference is reported back so it
appears in the log and in the workbook.

Accepted columns (case-insensitive, several spellings each, Chinese headers included)

    frequency    frequency, freq, f, f_hz, hz, 频率                       REQUIRED
    Z real       zreal, z_re, re, zr, real, z', 实部                      REQUIRED (or |Z|+phase)
    Z imag       zimag, z_im, im, zi, imag, z'', 虚部                     REQUIRED (or |Z|+phase)
    -Z imag      -zim, -z'', zim_neg, negzim, minus_zim                  sign flipped on read
    |Z|, phase   zmag/zabs/modulus + phase/theta/zphase (deg or rad)     polar alternative
    temperature  temperature, temp, t, t_degc, t_k, 温度                  optional, default 25 degC
    SOC          soc, state_of_charge, 荷电状态                            optional, default 0.5
    cell         cell, cell_id, sample, id, 电池                          optional, groups conditions

A headerless numeric file is accepted too: three columns are read as f, Z', Z''; five as
f, Z', Z'', T, SOC.

Spectra are separated by the cell/condition columns, or — with `split="auto"` — wherever
the frequency sequence restarts, which is what lets several spectra be concatenated head
to tail in one file. Conditions measured on different frequency vectors are resampled
onto a common grid (log-log interpolation over the shared range), because the stacked
forward model evaluates one frequency block per condition.
"""
from __future__ import annotations

import csv
import re
from collections import OrderedDict

import numpy as np

DEFAULT_TEMPERATURE_K = 298.15
DEFAULT_SOC = 0.5

ALIASES = {
    "freq": ("frequency", "freq", "f", "f_hz", "hz", "frequency_hz", "freq_hz",
             "w", "omega", "频率", "频率_hz"),
    "zre": ("zre", "zreal", "z_re", "z_real", "re", "zr", "real", "z1", "z'",
            "zprime", "re_z", "rez", "z_1", "实部", "阻抗实部"),
    "zim": ("zim", "zimag", "z_im", "z_imag", "z_imaginary", "im", "zi", "imag",
            "z2", "z''", "zdoubleprime", "im_z", "imz", "z_2", "虚部", "阻抗虚部"),
    "zim_neg": ("-zim", "-z''", "-z_im", "-im", "-zimag", "-z_imag", "-imz", "-z2",
                "zim_neg", "negzim", "neg_zim", "minus_zim", "z''_neg", "负虚部"),
    "zmag": ("zmag", "zabs", "z_mag", "modulus", "z_modulus", "absz", "mag", "|z|", "模"),
    "phase": ("phase", "theta", "zphase", "phase_deg", "phase_rad", "phi",
              "angle", "相位"),
    "temp": ("temperature", "temp", "t", "t_degc", "tdegc", "t_k", "tk",
             "temperature_k", "temperature_c", "温度"),
    "soc": ("soc", "state_of_charge", "soc_pct", "soc_percent", "荷电状态"),
    "cell": ("cell", "cell_id", "cellid", "sample", "sample_id", "id", "battery",
             "电池", "电池编号"),
}


_UNITS = {"ohm", "ohms", "ω", "Ω", "mohm", "kohm", "hz", "khz", "mhz", "s", "sec",
          "deg", "degc", "degf", "degree", "degrees", "celsius", "kelvin",
          "rad", "radians", "%", "pct", "percent",
          "v", "a", "k", "c", "℃", "°c", "欧", "欧姆", "赫兹"}
_STRIP = re.compile(r"[\(\[\{].*?[\)\]\}]")


def _norm(name: str) -> str:
    """Header -> comparison key: units in brackets and trailing unit tokens removed."""
    s = _STRIP.sub("", name.replace("﻿", "")).strip().lower()
    s = s.replace("/", "_").replace(",", "_").replace(" ", "_")
    parts = [p for p in s.split("_") if p]
    while len(parts) > 1 and parts[-1] in _UNITS:
        parts.pop()
    return "_".join(parts)


def _resolve_columns(header):
    norm = [_norm(h) for h in header]
    found = {}
    for key, names in ALIASES.items():
        for i, h in enumerate(norm):
            if h in names:
                found.setdefault(key, i)
    return found


class CsvFormatError(ValueError):
    """Raised when the CSV cannot be interpreted; carries the header for the message."""

    def __init__(self, header):
        super().__init__("unusable CSV header")
        self.header = header


def probe(path, group_column=None, max_rows=400):
    """Report what a file supplies, reading only its head.

    The interface calls this the moment a file is chosen, so the panel can show which
    quantities come from the file itself and disable the fallbacks for those that do.
    Never raises: an unreadable file simply reports nothing found.
    """
    out = dict(header=[], found={}, polar=False, negated=False, celsius=None,
               soc_percent=None, n_columns=0, headerless=False, error="")
    try:
        with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
            sample = fh.read(8192)
            fh.seek(0)
            try:
                reader = csv.reader(fh, csv.Sniffer().sniff(sample, delimiters=",;\t| "))
            except csv.Error:
                reader = csv.reader(fh)
            rows = []
            for r in reader:
                if any(c.strip() for c in r):
                    rows.append(r)
                if len(rows) >= max_rows:
                    break
    except OSError as e:
        out["error"] = str(e)
        return out
    if not rows:
        return out

    header = rows[0]
    out["header"] = header
    out["n_columns"] = len(header)
    if _is_numeric_row(header):
        out["headerless"] = True
        cols = ({"freq": 0, "zre": 1, "zim": 2, "temp": 3, "soc": 4} if len(header) >= 5
                else {"freq": 0, "zre": 1, "zim": 2} if len(header) >= 3 else {})
        body = rows
    else:
        cols = _resolve_columns(header)
        if group_column:
            gn = _norm(group_column)
            for i, h in enumerate(header):
                if _norm(h) == gn:
                    cols["cell"] = i
                    break
        body = rows[1:]

    out["polar"] = "zre" not in cols and "zmag" in cols and "phase" in cols
    out["negated"] = "zim" not in cols and "zim_neg" in cols
    for key in ("freq", "zre", "zim", "temp", "soc", "cell"):
        out["found"][key] = key in cols
    if out["polar"]:
        out["found"]["zre"] = out["found"]["zim"] = True

    def column(key):
        vals = []
        for r in body:
            try:
                vals.append(float(r[cols[key]]))
            except (ValueError, IndexError, KeyError):
                continue
        return vals

    if cols.get("temp") is not None and out["found"].get("temp"):
        v = column("temp")
        if v:
            out["celsius"] = max(v) < 200.0
    if cols.get("soc") is not None and out["found"].get("soc"):
        v = column("soc")
        if v:
            out["soc_percent"] = max(v) > 1.5
    return out


def _is_numeric_row(row):
    ok = 0
    for c in row:
        try:
            float(c)
            ok += 1
        except ValueError:
            return False
    return ok >= 3


def read_spectra(path, split="auto", group_column=None, negate_imag=False,
                 default_temperature_K=DEFAULT_TEMPERATURE_K, default_soc=DEFAULT_SOC):
    """Parse the CSV into cells -> conditions -> spectrum arrays.

    Returns `(cells, meta)`. `cells` maps cell id -> list of dicts
    ``{temperature_K, soc, freq, z}`` in the order encountered. `meta` reports every
    inference the reader made: `warnings` holds i18n keys with format arguments, so the
    caller decides how to phrase them.
    """
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t| ")
            reader = csv.reader(fh, dialect)
        except csv.Error:
            reader = csv.reader(fh)
        rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        raise CsvFormatError([])

    warnings = []
    header = rows[0]
    body = rows[1:]

    if _is_numeric_row(header):                       # headerless numeric file
        n = len(header)
        if n >= 5:
            cols = {"freq": 0, "zre": 1, "zim": 2, "temp": 3, "soc": 4}
        elif n >= 3:
            cols = {"freq": 0, "zre": 1, "zim": 2}
        else:
            raise CsvFormatError(header)
        body = rows
        warnings.append(("warn_headerless", {"n": n}))
    else:
        cols = _resolve_columns(header)

    if group_column:
        gn = _norm(group_column)
        for i, h in enumerate(header):
            if _norm(h) == gn:
                cols["cell"] = i
                break

    polar = "zre" not in cols and "zmag" in cols and "phase" in cols
    flip = negate_imag
    if "zim" not in cols and "zim_neg" in cols:
        cols["zim"] = cols["zim_neg"]
        flip = not flip
        warnings.append(("warn_neg_imag", {}))
    if "freq" not in cols or not (polar or ("zre" in cols and "zim" in cols)):
        raise CsvFormatError(header)

    phase_deg = True
    if polar:
        warnings.append(("warn_polar", {}))

    records = []
    for r in body:
        try:
            f = float(r[cols["freq"]])
            if polar:
                mag = float(r[cols["zmag"]])
                ph = float(r[cols["phase"]])
                records.append(dict(f=f, mag=mag, ph=ph))
                continue
            zr = float(r[cols["zre"]])
            zi = float(r[cols["zim"]])
        except (ValueError, IndexError):
            continue                       # blank or garbled row: skip silently
        rec = dict(f=f, zr=zr, zi=-zi if flip else zi)
        try:
            rec["t"] = float(r[cols["temp"]]) if "temp" in cols else None
        except (ValueError, IndexError):
            rec["t"] = None
        try:
            rec["soc"] = float(r[cols["soc"]]) if "soc" in cols else None
        except (ValueError, IndexError):
            rec["soc"] = None
        rec["cell"] = (r[cols["cell"]].strip()
                       if "cell" in cols and cols["cell"] < len(r) else "")
        records.append(rec)

    if polar:                              # second pass: rebuild the rectangular form
        phs = [r["ph"] for r in records]
        phase_deg = max(abs(p) for p in phs) > 3.2 if phs else True
        conv = (np.pi / 180.0) if phase_deg else 1.0
        rebuilt = []
        for r, src in zip(records, body):
            z = r["mag"] * np.exp(1j * r["ph"] * conv)
            rec = dict(f=r["f"], zr=float(z.real),
                       zi=float(-z.imag if flip else z.imag), t=None, soc=None, cell="")
            for key, fld in (("temp", "t"), ("soc", "soc")):
                if key in cols:
                    try:
                        rec[fld] = float(src[cols[key]])
                    except (ValueError, IndexError):
                        pass
            if "cell" in cols and cols["cell"] < len(src):
                rec["cell"] = src[cols["cell"]].strip()
            rebuilt.append(rec)
        records = rebuilt

    records = [r for r in records if np.isfinite(r["f"]) and r["f"] > 0]
    if not records:
        raise CsvFormatError(header)

    # temperature: values below 200 are read as Celsius; absent means room temperature
    if "temp" not in cols or all(r["t"] is None for r in records):
        for r in records:
            r["t_k"] = default_temperature_K
        warnings.append(("warn_default_temp", {"t": default_temperature_K - 273.15}))
        celsius = False
    else:
        t_vals = [r["t"] for r in records if r["t"] is not None]
        celsius = max(t_vals) < 200.0
        for r in records:
            base = r["t"] if r["t"] is not None else (
                default_temperature_K - 273.15 if celsius else default_temperature_K)
            r["t_k"] = base + 273.15 if celsius else base

    # SOC: values above 1.5 are read as percent; absent means mid-charge
    if "soc" not in cols or all(r["soc"] is None for r in records):
        for r in records:
            r["soc"] = default_soc
        warnings.append(("warn_default_soc", {"s": default_soc}))
    else:
        vals = [r["soc"] for r in records if r["soc"] is not None]
        scale = 100.0 if max(vals) > 1.5 else 1.0
        for r in records:
            r["soc"] = (r["soc"] / scale) if r["soc"] is not None else default_soc
        for r in records:
            r["soc"] = float(np.clip(r["soc"], 0.0, 1.0))

    # ---- segment into spectra
    segments, cur = [], [records[0]]
    for prev, rec in zip(records, records[1:]):
        new_seg = (rec["cell"] != prev["cell"]
                   or abs(rec["t_k"] - prev["t_k"]) > 1e-9
                   or abs(rec["soc"] - prev["soc"]) > 1e-9)
        if not new_seg and split == "auto" and len(cur) >= 2:
            descending = cur[1]["f"] < cur[0]["f"]
            if descending and rec["f"] > prev["f"] * 1.5:
                new_seg = True
            if (not descending) and rec["f"] < prev["f"] / 1.5:
                new_seg = True
        if new_seg:
            segments.append(cur)
            cur = []
        cur.append(rec)
    segments.append(cur)

    cells = OrderedDict()
    for seg in segments:
        cid = seg[0]["cell"] or "cell-1"
        order = np.argsort([r["f"] for r in seg])          # ascending, model-independent
        cells.setdefault(cid, []).append(dict(
            temperature_K=float(np.mean([r["t_k"] for r in seg])),
            soc=float(np.mean([r["soc"] for r in seg])),
            freq=np.array([seg[i]["f"] for i in order], float),
            z=np.array([seg[i]["zr"] + 1j * seg[i]["zi"] for i in order], complex)))

    # ---- one frequency vector per cell (the stacked model needs equal blocks)
    for cid, spectra in cells.items():
        n_before = len(spectra[0]["freq"])
        if _harmonise(spectra):
            warnings.append(("warn_resampled",
                             {"cell": cid, "before": n_before,
                              "after": len(spectra[0]["freq"])}))

    meta = dict(rows=len(records), spectra=len(segments), cells=len(cells),
                celsius_input=celsius, polar_input=polar, negated=flip,
                warnings=warnings)
    return cells, meta


def _harmonise(spectra, rtol=1e-3) -> bool:
    """Put one cell's conditions on a common frequency vector. True if data was resampled.

    Identical vectors (the normal case) are left untouched — no interpolation error is
    introduced unless the measurement itself used different frequencies per condition.
    """
    ref = spectra[0]["freq"]
    if all(s["freq"].size == ref.size and np.allclose(s["freq"], ref, rtol=rtol)
           for s in spectra):
        return False
    lo = max(s["freq"].min() for s in spectra)
    hi = min(s["freq"].max() for s in spectra)
    if not (hi > lo):
        raise ValueError("the spectra of one cell share no overlapping frequency range")
    n = int(np.median([s["freq"].size for s in spectra]))
    grid = np.logspace(np.log10(lo), np.log10(hi), max(n, 8))
    for s in spectra:
        lf, lg = np.log10(s["freq"]), np.log10(grid)
        s["z"] = (np.interp(lg, lf, s["z"].real)
                  + 1j * np.interp(lg, lf, s["z"].imag))
        s["freq"] = grid
    return True


def grid_of(cell_spectra):
    """Conditions and frequency vector implied by one cell's spectra."""
    conditions = [(s["temperature_K"], s["soc"]) for s in cell_spectra]
    return conditions, cell_spectra[0]["freq"]


def stack(cell_spectra) -> np.ndarray:
    """Concatenate the per-condition spectra into the model's stacked observation."""
    return np.concatenate([s["z"] for s in cell_spectra])


def compare_grids(conditions, freq, ref_conditions, ref_freq, rtol=1e-3) -> dict:
    """Axis-by-axis comparison against the grid a frozen system was calibrated on.

    A single in/out verdict is too blunt to act on: a file may carry exactly the same
    information budget and differ only in, say, the temperature span. The report names
    which axis moved, so the operator can judge whether re-freezing on this design is
    worth it rather than being told only that something is different.
    """
    temps = sorted({t for t, _ in conditions})
    socs = sorted({s for _, s in conditions})
    ref_temps = sorted({t for t, _ in ref_conditions})
    ref_socs = sorted({s for _, s in ref_conditions})

    def same(a, b, tol):
        return len(a) == len(b) and all(abs(x - y) <= tol for x, y in zip(a, b))

    rep = dict(
        n_conditions=len(conditions), ref_n_conditions=len(ref_conditions),
        n_frequencies=len(freq), ref_n_frequencies=len(ref_freq),
        temperatures=temps, ref_temperatures=ref_temps,
        socs=socs, ref_socs=ref_socs,
        temperatures_match=same(temps, ref_temps, 0.5),
        socs_match=same(socs, ref_socs, 0.02),
        frequencies_match=(len(freq) == len(ref_freq)
                           and bool(np.allclose(np.sort(freq), np.sort(ref_freq), rtol=rtol))),
        size_match=(len(conditions) == len(ref_conditions) and len(freq) == len(ref_freq)),
    )
    rep["conditions_match"] = (rep["temperatures_match"] and rep["socs_match"]
                              and len(conditions) == len(ref_conditions))
    rep["match"] = rep["conditions_match"] and rep["frequencies_match"]
    # same point count, different placement: the information budget is unchanged but the
    # design is not the calibrated one, which is the case worth calling out separately
    rep["same_size_different_design"] = rep["size_match"] and not rep["match"]
    return rep


def grids_match(conditions, freq, ref_conditions, ref_freq, rtol=1e-3) -> bool:
    """Does this acquisition grid match the one the frozen system was calibrated on?"""
    return compare_grids(conditions, freq, ref_conditions, ref_freq, rtol)["match"]


def describe_axis(values, kelvin=False) -> str:
    """Compact rendering of a grid axis for the log: '-10 / 5 / 20 / 35 / 50'."""
    if kelvin:
        values = [v - 273.15 for v in values]
    if len(values) > 6:
        return f"{values[0]:.4g} … {values[-1]:.4g} ({len(values)})"
    return " / ".join(f"{v:.4g}" for v in values)


# ----------------------------------------------------------------------------------
# XLSX export
# ----------------------------------------------------------------------------------
def write_xlsx(path, results, summary, log_lines, T):
    """Write the workbook. `T` is the translation callable so sheets follow the UI language.

    results: list of dicts {cell, coord, family, estimate, p, threshold, arm, certified,
    in_calibration}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    head_font = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="DDE5F0")
    warn_fill = PatternFill("solid", fgColor="FFE0E0")

    def sheet(title, headers):
        ws = wb.create_sheet(title=title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.font = head_font
            c.fill = head_fill
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)

    cols = [T("col_cell"), T("col_param"), T("col_family"), T("col_value"),
            T("col_prob"), T("col_thr"), T("col_arm"), T("col_status")]

    ws = sheet(T("sheet_claims"), cols)
    for r in results:
        if not r["certified"]:
            continue
        ws.append([r["cell"], r["coord"], r["family"], r["estimate"], r["p"],
                   r["threshold"], r["arm"],
                   T("val_in_cal") if r["in_calibration"] else T("val_out_cal")])
        if not r["in_calibration"]:
            ws.cell(ws.max_row, 8).fill = warn_fill

    ws = sheet(T("sheet_all"), cols[:7] + [T("col_certified"), T("col_status")])
    for r in results:
        ws.append([r["cell"], r["coord"], r["family"], r["estimate"], r["p"],
                   r["threshold"], r["arm"],
                   T("val_yes") if r["certified"] else T("val_no"),
                   T("val_in_cal") if r["in_calibration"] else T("val_out_cal")])

    ws = sheet(T("sheet_summary"), [T("sum_key"), T("sum_val")])
    for k, v in summary:
        ws.append([k, v])

    ws = sheet(T("sheet_log"), ["#", T("sheet_log")])
    for i, line in enumerate(log_lines, 1):
        ws.append([i, line])

    for s in wb.worksheets:
        for col in s.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            s.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    wb.save(path)
    return path
